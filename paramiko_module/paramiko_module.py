#!/usr/bin/env python
# !!!! dikkat snmp ve aaa configleri "]","$" karekterlerini bulundurmaktadır.
import paramiko
import time
import openpyxl

class para:
    def __init__(self, woe):
        self.woe = woe
        self.endswith = (b"$", b">", b"]", b"name:", b"word:")
        if woe == "west":
            s1, server1 = self.slave_server("10.***.***.116", username="username", password="password")
            s2, server2 = self.slave_server("10.***.***.146", username="username", password="password")
            print("Memory of.116 " + str(s1) + ", " + "Memory of.146 " + str(s2))
            if s1 > s2:
                print("connecting 10.***.***.146 \n.................................................")
                server1.close()
                self.server = server2
            elif s2 > s1:
                print("connecting 10.***.***.116 \n.................................................")
                server2.close()
                self.server = server1
        elif woe == "east":
            s1, server1 = self.slave_server("10.***.***.220", username="username", password="password")
            s2, server2 = self.slave_server("10.***.***.230", username="username", password="password")
            print("Memory of.220 " + str(s1) + ", " + "Memory of.230 " + str(s2))
            if s1 > s2:
                print("connecting 10.***.***.230 \n.................................................")
                server1.close()
                self.server = server2
            elif s2 > s1:
                print("connecting 10.***.***.220 \n.................................................")
                server2.close()
                self.server = server1

        self.shell = self.server.invoke_shell()
        self.out = self.shell.recv(4096)
        self.last = self.out
        while not self.last.endswith("$".encode("utf-8")):
            time.sleep(0.1)
            self.last = self.shell.recv(4096)
            self.out += self.last


    def slave_server(self,ip, username, password):
        client = paramiko.client.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(ip, username=username, password=password, look_for_keys=False)

        if ip == "10.***.***.116" or ip == "10.***.***.146" or ip == "10.***.***.220" or ip == "10.176.254.230":
            _, stdout, _ = client.exec_command(""" free -g | grep -i Mem | tr -s ' ' | cut -d' ' -f3 """)
            time.sleep(1)
            data = stdout.channel.recv(4096)
            return float(str(data, "utf-8")), client
        elif ip == "10.20.30.40": # this part find cpu usage of solaris server
            _, stdout, _ = client.exec_command("prstat")
            time.sleep(1)
            data = str(stdout.channel.recv(4096))
            data = data[data.rfind(":") + 1:-3]
            sum = 0
            for i in data.split(","): sum += float(i)
            return sum, client

    def ne_login(self, ip):
        self.out = self.last
        self.successfull = False
        self.command("telnet " + ip, 5)
        self.name = ip
        if self.last.endswith(b"name:"):
            self.command("username")
            self.command("password*")
            if b">" in self.last:
                self.successfull = True
                self.name = self.last[self.last .find("<".encode("utf-8")) + 1:-1].decode("utf-8")
            else:
                self.command("\x1A")
        else:
            self.command("\x1A")

    def command(self, com, timer=None):
        if timer == None:
            self.shell.send(com + "\n")
            self.last = self.shell.recv(4096)
            self.out += self.last
            while (not self.last.endswith(self.endswith)):
                time.sleep(0.1)
                self.last = self.shell.recv(4096)
                self.out += self.last
        else:
            self.shell.send(com + "\n")
            time.sleep(timer)
            self.last = self.shell.recv(4096)
            self.out += self.last

    def all_close(self):
        self.server.close()

    def issuccessfull(self):
        return self.successfull

    def ne_name(self):
        return self.name

    def printout(self):
        print("---------------------" + self.woe + "-----------------------------------------------------")
        print(str(self.out, "utf-8").replace("\r", ""))


def excel_page(way, sheet_name=None):
    wb = openpyxl.load_workbook(filename=way)
    if sheet_name == None:
        longexcel = list()
        for sheet in wb:
            excel = list()
            for row in sheet.iter_rows():
                line = []
                for cell in row:
                    line.append(cell.internal_value)
                excel.append(line)
            longexcel.append(excel)
        return longexcel, len(wb._sheets)
    else:
        excel = list()
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            line = []
            for cell in row:
                line.append(cell.internal_value)
            excel.append(line)
        return excel, len(wb._sheets)

if __name__ == "__main__":
    print("")